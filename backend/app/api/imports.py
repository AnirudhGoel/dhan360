"""Import endpoints: file uploads (CSV / CAS JSON / CAS PDF), manual entry, and history."""

from __future__ import annotations

import csv
import io
import json

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import desc, select
from sqlalchemy.orm import Session

from app.db.database import get_session
from app.db.models import ImportBatch
from app.parsers import cas_json, cas_pdf, generic_csv, zerodha_holdings, zerodha_tradebook
from app.parsers.manual import ManualEntry, parse_entries
from app.services.import_service import process_parse_result

router = APIRouter(prefix="/api/imports", tags=["imports"])

_CSV_PARSERS = {
    "zerodha_holdings": zerodha_holdings.parse,
    "zerodha_tradebook": zerodha_tradebook.parse,
    "generic_csv": generic_csv.parse,
}


def _to_csv_text(raw: bytes, filename: str) -> str:
    """Decode an upload to CSV text — converting .xlsx/.xls via openpyxl so users needn't."""
    if filename.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook  # lazy: only needed for spreadsheet uploads

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return ""
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in row])
        return buf.getvalue()
    return raw.decode("utf-8-sig", errors="replace")


def _serialize_batch(b: ImportBatch) -> dict:
    return {
        "id": b.id, "source": b.source, "file_name": b.file_name, "status": b.status,
        "count_parsed": b.count_parsed, "count_imported": b.count_imported,
        "count_merged": b.count_merged, "count_duplicate": b.count_duplicate,
        "count_skipped": b.count_skipped, "count_unclassified": b.count_unclassified,
        "diagnostics": json.loads(b.diagnostics) if b.diagnostics else [],
        "created_at": b.created_at.isoformat(),
    }


@router.get("")
def list_imports(db: Session = Depends(get_session)) -> list[dict]:
    batches = db.scalars(select(ImportBatch).order_by(desc(ImportBatch.created_at))).all()
    return [_serialize_batch(b) for b in batches]


@router.post("/upload")
async def upload(
    source: str = Form(...),
    account_name: str | None = Form(None),
    password: str | None = Form(None),
    file: list[UploadFile] = File(...),
    db: Session = Depends(get_session),
) -> dict:
    if not file:
        raise HTTPException(400, "No file uploaded.")
    first = file[0]
    raw = await first.read()
    fname = first.filename or source
    kwargs = {"account_name": account_name} if account_name else {}

    if source == "cas_pdf":
        result = cas_pdf.parse_bytes(raw, password or "", file_name=fname)
    elif source == "cas_json":
        try:
            data = json.loads(raw.decode("utf-8"))
        except (ValueError, UnicodeDecodeError) as exc:
            raise HTTPException(400, f"Invalid JSON: {exc}") from exc
        result = cas_json.parse_dict(data, file_name=fname)
    elif source == "zerodha_tradebook":
        # Several yearly exports may be selected at once — combine and de-duplicate them.
        contents = [_to_csv_text(raw, fname)]
        for extra in file[1:]:
            contents.append(_to_csv_text(await extra.read(), extra.filename or source))
        name = f"{len(file)} tradebook files" if len(file) > 1 else fname
        result = zerodha_tradebook.parse(contents, file_name=name, **kwargs)
    elif source in _CSV_PARSERS:
        result = _CSV_PARSERS[source](_to_csv_text(raw, fname), file_name=fname, **kwargs)
    else:
        raise HTTPException(400, f"Unknown source '{source}'.")

    batch = process_parse_result(db, result)
    db.commit()
    return _serialize_batch(batch)


@router.post("/manual")
def manual(entries: list[ManualEntry], db: Session = Depends(get_session)) -> dict:
    if not entries:
        raise HTTPException(400, "No entries provided.")
    result = parse_entries(entries, file_name="manual-entry")
    batch = process_parse_result(db, result)
    db.commit()
    return _serialize_batch(batch)
